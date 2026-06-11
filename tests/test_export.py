from pathlib import Path

from openpyxl import load_workbook

from bay_area_projectintel.db import Database
from bay_area_projectintel.export.excel import DEADLINE_FILL, HEADERS, export_excel
from bay_area_projectintel.models import Category, Company, Project


def test_export_hard_filters_contact_rows(tmp_path: Path) -> None:
    db = Database(tmp_path / "test.sqlite3")
    db.migrate()

    with_contact = Project(
        raw_record_id=_raw(db, "1"),
        source="datasf-building-permits",
        source_record_id="1",
        description="restaurant remodel",
        project_date="2026-06-01",
        company=Company(name="Acme Builders", email="info@example.com"),
        category=Category.RESTAURANT_RETAIL,
        confidence=0.9,
        content_hash="p1",
    )
    pending = Project(
        raw_record_id=_raw(db, "2"),
        source="datasf-building-permits",
        source_record_id="2",
        description="office remodel",
        project_date="2026-06-01",
        company=Company(name="No Contact Builders"),
        category=Category.OFFICE_LAB,
        confidence=0.9,
        content_hash="p2",
    )
    old = Project(
        raw_record_id=_raw(db, "3"),
        source="datasf-building-permits",
        source_record_id="3",
        description="old restaurant remodel",
        project_date="2020-01-01",
        company=Company(name="Old Builders", email="old@example.com"),
        category=Category.RESTAURANT_RETAIL,
        confidence=0.9,
        content_hash="p3",
    )
    db.upsert_project(with_contact)
    db.upsert_project(pending)
    db.upsert_project(old)

    out = tmp_path / "leads.xlsx"
    stats = export_excel(db, out)
    wb = load_workbook(out)

    assert stats["total"] == 2
    assert stats["exported"] == 1
    assert stats["pending"] == 1
    assert "Summary" not in wb.sheetnames
    assert "RESTAURANT_RETAIL" in wb.sheetnames
    assert "待补全 (Pending)" in wb.sheetnames
    assert wb["RESTAURANT_RETAIL"].cell(1, 1).value == "项目描述"
    assert "状态" not in [cell.value for cell in wb["RESTAURANT_RETAIL"][1]]
    assert wb["RESTAURANT_RETAIL"].max_row == 2
    assert wb["待补全 (Pending)"].max_row == 2


def test_export_highlights_bid_deadline(tmp_path: Path) -> None:
    db = Database(tmp_path / "test.sqlite3")
    db.migrate()
    rfp = Project(
        raw_record_id=_raw(db, "10"),
        source="samgov-construction",
        source_record_id="10",
        description="HVAC renovation RFP",
        bid_deadline="2026-06-15",
        company=Company(name="GSA Public Buildings", email="co@gsa.gov"),
        category=Category.PUBLIC_WORKS,
        confidence=0.9,
        content_hash="r1",
    )
    db.upsert_project(rfp)

    out = tmp_path / "leads.xlsx"
    export_excel(db, out)
    sheet = load_workbook(out)["PUBLIC_WORKS"]

    deadline_col = HEADERS.index("投标截止") + 1
    assert sheet.cell(1, deadline_col).value == "投标截止"
    assert sheet.cell(2, deadline_col).value == "2026-06-15"
    assert sheet.cell(2, deadline_col).fill.fgColor.rgb == DEADLINE_FILL.fgColor.rgb


def _raw(db: Database, source_record_id: str) -> int:
    from bay_area_projectintel.models import RawRecord

    payload = {"record_id": source_record_id, "description": "x"}
    row_id, _ = db.upsert_raw_record(
        RawRecord(
            source="datasf-building-permits",
            source_record_id=source_record_id,
            payload=payload,
            content_hash=source_record_id,
        )
    )
    return row_id
