from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from bay_area_projectintel.db import Database
from bay_area_projectintel.models import Category


HEADERS = ["项目描述", "城市", "日期", "投标截止", "公司名", "邮箱", "电话", "首次发现日期", "来源链接"]
DEADLINE_FILL = PatternFill("solid", fgColor="F4CCCC")
HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")
DEFAULT_EXPORT_LOOKBACK_DAYS = 90


def export_excel(db: Database, out: Path, category: str | None = None, lookback_days: int = DEFAULT_EXPORT_LOOKBACK_DAYS) -> dict[str, int]:
    rows = _recent_rows(db.export_rows(category=category), lookback_days=lookback_days)
    wb = Workbook()
    first_sheet = wb.active
    used_first_sheet = False

    def make_sheet(title: str) -> Any:
        nonlocal used_first_sheet
        if not used_first_sheet:
            first_sheet.title = title
            used_first_sheet = True
            return first_sheet
        return wb.create_sheet(title)

    by_category: dict[str, list[Any]] = defaultdict(list)
    pending: list[Any] = []
    for row in rows:
        has_contact = bool(row["email"] or row["phone"])
        if has_contact:
            by_category[row["category"] or Category.OTHER.value].append(row)
        else:
            pending.append(row)

    for category_name in Category:
        data = by_category.get(category_name.value, [])
        if data:
            _write_rows(make_sheet(category_name.value[:31]), data)

    pending_sheet = make_sheet("待补全 (Pending)")
    _write_rows(pending_sheet, pending)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return {"total": len(rows), "exported": sum(len(v) for v in by_category.values()), "pending": len(pending)}


def _recent_rows(rows: list[Any], lookback_days: int) -> list[Any]:
    cutoff = date.today() - timedelta(days=lookback_days)
    recent = []
    for row in rows:
        project_date = _parse_date(row["project_date"])
        if project_date is None or project_date >= cutoff:
            recent.append(row)
    return recent


def _parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    text = text.split("T", 1)[0].split(" ", 1)[0]
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    return None


def _write_rows(sheet: Any, rows: list[Any]) -> None:
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    deadline_col = HEADERS.index("投标截止") + 1
    for row in rows:
        bid_deadline = row["bid_deadline"]
        sheet.append(
            [
                row["description"],
                row["city"],
                row["project_date"],
                bid_deadline,
                row["company_name"],
                row["email"],
                row["phone"],
                str(row["first_seen"]).split("T", 1)[0],
                row["source_url"],
            ]
        )
        excel_row = sheet.max_row
        if bid_deadline:
            sheet.cell(excel_row, deadline_col).fill = DEADLINE_FILL
        link_cell = sheet.cell(excel_row, len(HEADERS))
        if link_cell.value:
            link_cell.hyperlink = link_cell.value
            link_cell.style = "Hyperlink"

    sheet.freeze_panes = "A2"
    _autosize(sheet)


def _autosize(sheet: Any) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 80))
        sheet.column_dimensions[column].width = max(12, min(max_length + 2, 60))
